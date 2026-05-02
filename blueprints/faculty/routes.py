import os
import re
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import (
    abort,
    current_app,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from sqlalchemy import bindparam, func, text
from sqlalchemy.exc import IntegrityError
from werkzeug.exceptions import Forbidden, Unauthorized
from werkzeug.security import check_password_hash, generate_password_hash

from app import cache, db
from models import (
    Bill,
    Budget,
    ExpensePrintReport,
    FacultyBudgetCycle,
    FacultyMessage,
    FacultyMessageFloor,
    FacultyReportSubmission,
    Feedback,
    Menu,
    Suggestion,
    SuggestionVote,
    User,
)
from . import faculty_bp
from ..budgeting import build_floor_budget_ledger
from ..utils import (
    _ensure_username_from_full_name,
    _get_active_floor,
    _get_tenant_floor_options,
    current_tenant_faculty_workflow_enabled,
    faculty_workflow_enabled_for_user,
    faculty_visible_users_query,
    log_tenant_audit,
    _require_faculty,
    _require_user,
    send_email_notification,
    send_push_notification,
    tenant_filter,
    visible_budget_condition,
)


@faculty_bp.before_request
def _faculty_auth_guard():
    shared_staff_endpoints = {
        'faculty.reports_page', 
        'faculty.download_floor_submission',
        'faculty.download_adhoc_report',
        'faculty.delete_adhoc_report'
    }
    if request.endpoint == 'faculty.login' or request.endpoint in shared_staff_endpoints:
        return None

    user = _require_user()
    if not user:
        session.clear()
        flash('Your Faculty session expired. Please sign in again.', 'error')
        return redirect(url_for('faculty.login'))

    if not getattr(user, 'is_active', True):
        session.clear()
        flash('This Faculty account is inactive. Please contact support.', 'error')
        return redirect(url_for('faculty.login'))

    if not faculty_workflow_enabled_for_user(user):
        session.clear()
        flash('Faculty workflow is disabled for this tenant right now.', 'error')
        return redirect(url_for('faculty.login'))

    if user.role != 'faculty':
        flash('Please sign in with a Faculty account to access the Faculty portal.', 'error')
        return redirect(url_for('faculty.login'))

    return None


@faculty_bp.errorhandler(Unauthorized)
@faculty_bp.errorhandler(401)
def _faculty_unauthorized(_error):
    session.clear()
    flash('Your Faculty session expired. Please sign in again.', 'error')
    return redirect(url_for('faculty.login'))


@faculty_bp.errorhandler(Forbidden)
@faculty_bp.errorhandler(403)
def _faculty_forbidden(_error):
    user = _require_user()
    if user and user.role == 'faculty':
        flash('You do not have permission to access that Faculty page.', 'error')
    else:
        session.clear()
        flash('Please sign in with a Faculty account to access the Faculty portal.', 'error')
    return redirect(url_for('faculty.login'))


def _tenant_slug():
    tenant_name = getattr(g, 'tenant_name', None) or 'tenant'
    slug = re.sub(r'[^a-z0-9]+', '-', tenant_name.lower()).strip('-')
    return slug or 'tenant'


def _report_storage_dir():
    base_dir = current_app.config["REPORT_STORAGE_ROOT"]
    path = os.path.join(base_dir, _tenant_slug())
    os.makedirs(path, exist_ok=True)
    return path


def _build_report_filename(cycle, floor, revision_no):
    return (
        f"{_tenant_slug()}_Floor-{floor}_"
        f"{cycle.start_date.strftime('%Y%m%d')}_to_{cycle.end_date.strftime('%Y%m%d')}"
        f"_v{revision_no}.pdf"
    )


def _current_active_cycle():
    if not current_tenant_faculty_workflow_enabled():
        return None
    return (
        tenant_filter(FacultyBudgetCycle.query)
        .filter_by(status='active')
        .order_by(FacultyBudgetCycle.start_date.desc())
        .first()
    )


def _cycle_for_floor(cycle_id, floor):
    return (
        tenant_filter(Budget.query)
        .filter(
            Budget.cycle_id == cycle_id,
            Budget.floor == floor,
            visible_budget_condition(),
        )
        .first()
    )


def _is_cycle_fully_verified(cycle):
    if not cycle:
        return True
    budgets = tenant_filter(Budget.query).filter_by(cycle_id=cycle.id).all()
    if not budgets:
        return True
    submissions = tenant_filter(FacultyReportSubmission.query).filter_by(cycle_id=cycle.id).all()
    submissions_by_floor = {s.floor: s for s in submissions}
    for budget in budgets:
        try:
            val = float(budget.amount_allocated)
        except (ValueError, TypeError):
            val = 0.0
        if val == 0.0:
            continue
        sub = submissions_by_floor.get(budget.floor)
        if not sub or sub.status != 'verified':
            return False
    return True


def _parse_selected_ids(values):
    parsed = []
    for raw in values:
        try:
            parsed.append(int(raw))
        except (TypeError, ValueError):
            continue
    return parsed


def _cycle_form_floor_options(user, cycle=None):
    floor_values = set(_get_tenant_floor_options(user))
    if cycle:
        existing_floors = (
            tenant_filter(Budget.query)
            .filter_by(cycle_id=cycle.id)
            .with_entities(Budget.floor)
            .all()
        )
        floor_values.update(floor for (floor,) in existing_floors)
    return sorted(floor_values)


def _parse_cycle_form(floor_options):
    title = (request.form.get('title') or '').strip()
    start_date_raw = request.form.get('start_date')
    end_date_raw = request.form.get('end_date')
    deadline_raw = request.form.get('submission_deadline')
    cycle_notes = (request.form.get('notes') or '').strip()

    if not title or not start_date_raw or not end_date_raw or not deadline_raw:
        return None, 'Cycle title, term dates, and deadline are required.'

    try:
        start_date = datetime.strptime(start_date_raw, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_raw, '%Y-%m-%d').date()
        submission_deadline = datetime.strptime(deadline_raw, '%Y-%m-%d').date()
    except ValueError:
        return None, 'Invalid cycle date provided.'

    if end_date < start_date:
        return None, 'Cycle end date must be after the start date.'

    if submission_deadline < start_date:
        return None, 'Submission deadline cannot be before the cycle starts.'

    allocations = []
    for floor in floor_options:
        amount_raw = request.form.get(f'amount_{floor}', '0')
        faculty_note = (request.form.get(f'faculty_note_{floor}') or '').strip()
        try:
            amount = float(amount_raw or 0)
        except ValueError:
            amount = 0

        allocations.append({
            'floor': floor,
            'amount': amount,
            'faculty_note': faculty_note,
        })

    return {
        'title': title,
        'start_date': start_date,
        'end_date': end_date,
        'submission_deadline': submission_deadline,
        'notes': cycle_notes,
        'allocations': allocations,
    }, None


def _sync_cycle_budgets(cycle, user, allocations):
    budget_note = cycle.notes or f'Faculty cycle {cycle.title}'
    existing_budgets = (
        tenant_filter(Budget.query)
        .filter_by(cycle_id=cycle.id)
        .all()
    )
    budgets_by_floor = {budget.floor: budget for budget in existing_budgets}

    for allocation in allocations:
        budget = budgets_by_floor.get(allocation['floor'])
        if not budget:
            budget = Budget(
                floor=allocation['floor'],
                cycle_id=cycle.id,
                tenant_id=getattr(g, 'tenant_id', None),
            )
            db.session.add(budget)

        budget.allocated_by_id = user.id
        budget.amount_allocated = allocation['amount']
        budget.allocation_type = 'faculty_cycle'
        budget.start_date = cycle.start_date
        budget.end_date = cycle.end_date
        budget.notes = budget_note
        budget.faculty_note = allocation['faculty_note']
        budget.is_faculty_allocation = True


def _submission_selectable_bills(floor, existing_submission=None):
    query = tenant_filter(Bill.query).filter(Bill.floor == floor)
    if existing_submission:
        query = query.filter(
            (Bill.report_submission_id.is_(None))
            | (Bill.report_submission_id == existing_submission.id)
        )
    else:
        query = query.filter(Bill.report_submission_id.is_(None))
    return query.order_by(Bill.bill_date.desc(), Bill.created_at.desc()).all()

def _saved_print_reports_for_floor(floor, cycle_id=None):
    query = tenant_filter(ExpensePrintReport.query).filter_by(floor=floor)
    if cycle_id:
        query = query.filter(ExpensePrintReport.cycle_id == cycle_id)
    return query.order_by(ExpensePrintReport.created_at.desc()).all()


def _save_submission_file(file_storage, cycle, floor, revision_no):
    filename = _build_report_filename(cycle, floor, revision_no)
    path = os.path.join(_report_storage_dir(), filename)
    file_storage.save(path)
    relative_path = f"{_tenant_slug()}/{filename}"
    return filename, relative_path, os.path.getsize(path)


def _safe_remove_file(path):
    if not path:
        current_app.logger.warning('Skipping PDF cleanup because no storage path was provided.')
        return
    try:
        if os.path.isabs(path):
            abs_path = path
        else:
            abs_path = os.path.normpath(os.path.join(current_app.config["REPORT_STORAGE_ROOT"], path))

        if os.path.exists(abs_path):
            os.remove(abs_path)
            current_app.logger.info('Deleted stored Faculty PDF during cleanup: %s', abs_path)
        else:
            current_app.logger.warning('Faculty PDF cleanup skipped because file was already missing: %s', abs_path)
    except OSError:
        current_app.logger.exception('Faculty PDF cleanup failed for path: %s', path)


def _message_target_floors(message):
    if message.target_scope == 'selected_floors':
        return sorted({target.floor for target in message.target_floors})
    return []


def _build_submission_verification_data(submission, linked_bills):
    allocation = _cycle_for_floor(submission.cycle_id, submission.floor)
    if submission.allocated_amount is not None:
        allocated_budget = float(submission.allocated_amount)
    else:
        allocated_budget = float(allocation.amount_allocated) if allocation else 0
    
    linked_bills_total = float(sum(float(bill.total_amount or 0) for bill in linked_bills))
    difference_amount = round(abs(allocated_budget - linked_bills_total), 2)
    verification_state = 'matched'
    if linked_bills_total < allocated_budget:
        verification_state = 'remaining'
    elif linked_bills_total > allocated_budget:
        verification_state = 'overspent'

    return {
        'allocation': allocation,
        'allocated_budget': allocated_budget,
        'linked_bills_total': linked_bills_total,
        'difference_amount': difference_amount,
        'verification_state': verification_state,
    }


def _delete_cycle_related_data(cycle):
    submissions = tenant_filter(FacultyReportSubmission.query).filter_by(cycle_id=cycle.id).all()
    for submission in submissions:
        for bill in list(submission.bills):
            bill.report_submission_id = None
        for expense in list(submission.expenses):
            expense.report_submission_id = None
        _safe_remove_file(submission.storage_path)
        db.session.delete(submission)

    print_reports = tenant_filter(ExpensePrintReport.query).filter_by(cycle_id=cycle.id).all()
    for report in print_reports:
        for link in list(report.bill_links):
            db.session.delete(link)
        db.session.delete(report)

    tenant_filter(Budget.query).filter_by(cycle_id=cycle.id).delete(synchronize_session=False)
    db.session.delete(cycle)


def _sync_submission_links(submission, selected_bills, selected_expenses):
    for bill in submission.bills:
        bill.report_submission_id = None
    for expense in submission.expenses:
        expense.report_submission_id = None

    for bill in selected_bills:
        bill.report_submission_id = submission.id
    for expense in selected_expenses:
        expense.report_submission_id = submission.id


FACULTY_IMPORT_HEADERS = ['TR', 'Name', 'Floor']
FACULTY_IMPORT_LIMIT = 500
FACULTY_MANAGED_ROLES = {'member', 'pantryHead', 'teaManager'}


def _clear_faculty_dashboard_cache(tenant_id=None):
    cache.delete_memoized(_get_faculty_dashboard_stats, tenant_id or getattr(g, 'tenant_id', None))


def _display_user_label(user):
    return user.full_name or user.username or user.email or f'User {user.id}'


def _global_existing_user_keys(tr_numbers, emails):
    tr_numbers = [str(tr) for tr in tr_numbers if tr]
    emails = [str(email) for email in emails if email]
    if not tr_numbers and not emails:
        return set(), set()

    stmt = text(
        'SELECT tr_number, email FROM "user" '
        'WHERE tr_number IN :tr_numbers OR email IN :emails'
    ).bindparams(
        bindparam('tr_numbers', expanding=True),
        bindparam('emails', expanding=True),
    )
    rows = db.session.execute(
        stmt,
        {
            'tr_numbers': tr_numbers or ['__none__'],
            'emails': emails or ['__none__'],
        },
    ).mappings()
    existing_trs = set()
    existing_emails = set()
    for row in rows:
        if row.get('tr_number'):
            existing_trs.add(str(row['tr_number']))
        if row.get('email'):
            existing_emails.add(str(row['email']).lower())
    return existing_trs, existing_emails


def _normalize_tr(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _normalize_floor(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _rows_from_workbook_upload(upload):
    if not upload or not upload.filename:
        return None, 'Please upload an Excel file.'
    if not upload.filename.lower().endswith('.xlsx'):
        return None, 'Only .xlsx files are supported.'

    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
    except Exception:
        return None, 'Unable to read the Excel file.'

    headers = [
        str(cell.value).strip() if cell.value is not None else ''
        for cell in next(sheet.iter_rows(min_row=1, max_row=1), [])
    ]
    headers = headers[:len(FACULTY_IMPORT_HEADERS)]
    if headers != FACULTY_IMPORT_HEADERS:
        return None, 'Invalid headers. Use exactly: TR, Name, Floor.'

    rows = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(cells[:3])
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append({
            'row_number': row_number,
            'tr': _normalize_tr(values[0] if len(values) > 0 else None),
            'name': (str(values[1]).strip() if len(values) > 1 and values[1] is not None else ''),
            'floor': _normalize_floor(values[2] if len(values) > 2 else None),
        })

    if not rows:
        return None, 'The Excel file has no member rows.'
    if len(rows) > FACULTY_IMPORT_LIMIT:
        return None, f'Import limit is {FACULTY_IMPORT_LIMIT} rows per file.'
    return rows, None


def _rows_from_import_payload(payload_rows):
    rows = []
    for index, row in enumerate(payload_rows or [], start=2):
        rows.append({
            'row_number': row.get('row_number') or index,
            'tr': _normalize_tr(row.get('tr') or row.get('TR')),
            'name': (row.get('name') or row.get('Name') or '').strip(),
            'floor': _normalize_floor(row.get('floor') or row.get('Floor')),
        })
    return rows


def _validate_import_rows(rows):
    from models import Tenant

    tenant = Tenant.query.get(getattr(g, 'tenant_id', None))
    floor_limit = tenant.floor_count if tenant and tenant.floor_count else 11
    seen_trs = set()
    tr_candidates = []
    email_candidates = []

    for row in rows:
        tr = row['tr']
        if tr:
            tr_candidates.append(tr)
            email_candidates.append(f'{tr}@jameasaifiyah.edu'.lower())

    existing_trs, existing_emails = _global_existing_user_keys(tr_candidates, email_candidates)
    valid_rows = []
    invalid_rows = []

    for row in rows:
        errors = []
        tr = row['tr']
        email = f'{tr}@jameasaifiyah.edu'.lower() if tr else ''

        if not re.fullmatch(r'\d{5}', tr or ''):
            errors.append('TR must be exactly 5 digits.')
        if tr in seen_trs:
            errors.append('Duplicate TR in this file.')
        if tr in existing_trs:
            errors.append('TR already exists.')
        if email and email in existing_emails:
            errors.append('Generated email already exists.')
        if row['floor'] is None:
            errors.append('Floor must be a number.')
        elif row['floor'] < 1 or row['floor'] > floor_limit:
            errors.append(f'Floor must be between 1 and {floor_limit}.')

        seen_trs.add(tr)
        prepared = {
            'row_number': row['row_number'],
            'tr': tr,
            'name': row['name'],
            'floor': row['floor'],
            'email': email,
        }
        if errors:
            prepared['errors'] = errors
            invalid_rows.append(prepared)
        else:
            valid_rows.append(prepared)

    return {
        'valid_rows': valid_rows,
        'invalid_rows': invalid_rows,
        'valid_count': len(valid_rows),
        'invalid_count': len(invalid_rows),
        'total_count': len(rows),
    }


@cache.memoize(timeout=300)
def _get_faculty_dashboard_stats(tenant_id):
    today = date.today()
    week_end = today + timedelta(days=6)
    first_of_month = date(today.year, today.month, 1)
    if today.month == 12:
        first_of_next_month = date(today.year + 1, 1, 1)
    else:
        first_of_next_month = date(today.year, today.month + 1, 1)

    population_roles = ['member', 'pantryHead', 'teaManager']
    user_filters = [
        User.tenant_id == tenant_id,
        User.is_active.is_(True),
        User.role.in_(population_roles),
    ]

    active_members = db.session.query(func.count(User.id)).filter(
        User.tenant_id == tenant_id,
        User.is_active.is_(True),
        User.role == 'member',
    ).scalar() or 0
    pantry_heads = db.session.query(func.count(User.id)).filter(
        User.tenant_id == tenant_id,
        User.is_active.is_(True),
        User.role == 'pantryHead',
    ).scalar() or 0
    tea_managers = db.session.query(func.count(User.id)).filter(
        User.tenant_id == tenant_id,
        User.is_active.is_(True),
        User.role == 'teaManager',
    ).scalar() or 0
    floor_counts = {
        int(floor): int(count)
        for floor, count in db.session.query(User.floor, func.count(User.id))
        .filter(*user_filters, User.floor.isnot(None))
        .group_by(User.floor)
        .order_by(User.floor.asc())
        .all()
    }
    planned_today = Menu.query.filter(Menu.tenant_id == tenant_id, Menu.date == today).count()
    planned_week = Menu.query.filter(Menu.tenant_id == tenant_id, Menu.date >= today, Menu.date <= week_end).count()
    planned_month = Menu.query.filter(
        Menu.tenant_id == tenant_id,
        Menu.date >= first_of_month,
        Menu.date < first_of_next_month,
    ).count()
    avg_rating = db.session.query(func.avg(Feedback.rating)).filter(
        Feedback.tenant_id == tenant_id,
        Feedback.menu_id.isnot(None),
    ).scalar() or 0

    return {
        'active_members': int(active_members),
        'pantry_heads': int(pantry_heads),
        'tea_managers': int(tea_managers),
        'floor_counts': floor_counts,
        'planned_today': int(planned_today),
        'planned_week': int(planned_week),
        'planned_month': int(planned_month),
        'avg_rating': round(float(avg_rating), 1),
    }


@faculty_bp.route('/faculty/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = (request.form.get('email') or '').strip()
        password = (request.form.get('password') or '').strip()

        if email and '@' not in email:
            email = f"{email}@jameasaifiyah.edu"

        user = User.query.filter_by(email=email, role='faculty').first()
        if not user or not getattr(user, 'is_active', True) or not user.password_hash or not check_password_hash(user.password_hash, password):
            flash('Invalid Faculty credentials', 'error')
            return render_template('faculty/login.html')
        if not faculty_workflow_enabled_for_user(user):
            flash('Faculty workflow is disabled for this tenant right now.', 'error')
            return render_template('faculty/login.html')

        if _ensure_username_from_full_name(user, db.session):
            db.session.commit()

        if user.is_first_login:
            session['temp_user_id'] = user.id
            return redirect(url_for('auth.change_password'))

        session['user_id'] = user.id
        session['role'] = user.role
        session['floor'] = user.floor
        session.permanent = True
        return redirect(url_for('faculty.dashboard'))

    return render_template('faculty/login.html')


@faculty_bp.route('/faculty/dashboard')
def dashboard():
    user = _require_faculty()
    today = date.today()
    tenant_id = getattr(g, 'tenant_id', None)
    stats = _get_faculty_dashboard_stats(tenant_id)
    active_cycle = _current_active_cycle()
    cycle_summary = {
        'pending_submission_count': 0,
        'verified_submission_count': 0,
        'overdue_floors': 0,
        'cycle_total_allocated': 0,
    }
    if active_cycle:
        cycle_allocations = (
            tenant_filter(Budget.query)
            .filter(Budget.cycle_id == active_cycle.id, visible_budget_condition())
            .order_by(Budget.floor.asc())
            .all()
        )
        submissions = tenant_filter(FacultyReportSubmission.query).filter_by(cycle_id=active_cycle.id).all()
        submissions_by_floor = {s.floor: s for s in submissions}
        cycle_summary['cycle_total_allocated'] = float(sum(float(b.amount_allocated) for b in cycle_allocations))
        for allocation in cycle_allocations:
            submission = submissions_by_floor.get(allocation.floor)
            if submission and submission.status == 'verified':
                cycle_summary['verified_submission_count'] += 1
            else:
                cycle_summary['pending_submission_count'] += 1
                if active_cycle.submission_deadline < today:
                    cycle_summary['overdue_floors'] += 1

    floor_rows = [
        {'floor': floor, 'count': stats['floor_counts'].get(floor, 0)}
        for floor in _get_tenant_floor_options(user)
    ]

    cycles = (
        tenant_filter(FacultyBudgetCycle.query)
        .order_by(FacultyBudgetCycle.created_at.desc())
        .limit(8)
        .all()
    )

    return render_template(
        'faculty/dashboard.html',
        current_user=user,
        stats=stats,
        active_cycle=active_cycle,
        cycle_summary=cycle_summary,
        floor_rows=floor_rows,
        cycles=cycles,
        today=today,
    )


@faculty_bp.route('/faculty/members')
def members():
    user = _require_faculty()
    users = (
        faculty_visible_users_query()
        .order_by(User.floor.asc(), User.role.asc(), User.tr_number.asc(), User.full_name.asc())
        .all()
    )
    role_counts = {
        'all': len(users),
        'pantryHead': len([u for u in users if u.role == 'pantryHead']),
        'teaManager': len([u for u in users if u.role == 'teaManager']),
    }
    return render_template(
        'faculty/members.html',
        current_user=user,
        users=users,
        role_counts=role_counts,
        floor_options=_get_tenant_floor_options(user),
    )


@faculty_bp.route('/faculty/members/<int:user_id>/role', methods=['POST'])
def update_member_role(user_id):
    user = _require_faculty()
    target = faculty_visible_users_query().filter(User.id == user_id).first_or_404()
    if target.role in {'admin', 'super_admin', 'faculty'}:
        abort(403)

    new_role = (request.form.get('role') or '').strip()
    if new_role not in FACULTY_MANAGED_ROLES:
        flash('Invalid role selected.', 'error')
        return redirect(url_for('faculty.members'))
    if target.role == new_role:
        flash('No role change was needed.', 'success')
        return redirect(url_for('faculty.members'))

    old_role = target.role
    target.role = new_role
    log_tenant_audit(
        'faculty_role_change',
        target_type='user',
        target_id=target.id,
        description=f'Changed {_display_user_label(target)} from {old_role} to {new_role}.',
        details={'old_role': old_role, 'new_role': new_role, 'floor': target.floor},
        actor_user=user,
    )
    db.session.commit()
    _clear_faculty_dashboard_cache()
    flash('Role updated successfully.', 'success')
    return redirect(url_for('faculty.members'))


@faculty_bp.route('/faculty/members/<int:user_id>/deactivate', methods=['POST'])
def deactivate_member(user_id):
    user = _require_faculty()
    target = faculty_visible_users_query().filter(User.id == user_id).first_or_404()
    if target.role in {'admin', 'super_admin', 'faculty'}:
        abort(403)
    if target.id == user.id:
        abort(403)

    target.is_active = False
    log_tenant_audit(
        'faculty_deactivate_user',
        target_type='user',
        target_id=target.id,
        description=f'Deactivated {_display_user_label(target)}.',
        details={'role': target.role, 'floor': target.floor, 'tr_number': target.tr_number},
        actor_user=user,
    )
    db.session.commit()
    _clear_faculty_dashboard_cache()
    flash('User deactivated successfully.', 'success')
    return redirect(url_for('faculty.members'))


@faculty_bp.route('/faculty/import/template')
def import_template():
    _require_faculty()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Members'
    sheet.append(FACULTY_IMPORT_HEADERS)
    sheet.append(['25687', 'Example Member', 1])
    for column, width in {'A': 14, 'B': 28, 'C': 12}.items():
        sheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'{_tenant_slug()}_faculty_import_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@faculty_bp.route('/faculty/import/validate', methods=['POST'])
def validate_import():
    _require_faculty()
    rows, error = _rows_from_workbook_upload(request.files.get('file'))
    if error:
        return jsonify({'error': error}), 400
    return jsonify(_validate_import_rows(rows))


@faculty_bp.route('/faculty/import/commit', methods=['POST'])
def commit_import():
    user = _require_faculty()
    payload = request.get_json(silent=True) or {}
    rows = _rows_from_import_payload(payload.get('rows') or [])
    if not rows:
        return jsonify({'error': 'No valid rows were submitted.'}), 400
    if len(rows) > FACULTY_IMPORT_LIMIT:
        return jsonify({'error': f'Import limit is {FACULTY_IMPORT_LIMIT} rows per commit.'}), 400

    report = _validate_import_rows(rows)
    users_to_create = []
    for row in report['valid_rows']:
        users_to_create.append(User(
            role='member',
            floor=row['floor'],
            tr_number=row['tr'],
            full_name=row['name'] or None,
            email=row['email'],
            password_hash=generate_password_hash('maskan1447'),
            is_first_login=True,
            is_verified=False,
            is_active=True,
            tenant_id=getattr(g, 'tenant_id', None),
        ))

    imported_count = 0
    if users_to_create:
        try:
            db.session.bulk_save_objects(users_to_create)
            imported_count = len(users_to_create)
        except IntegrityError:
            db.session.rollback()
            return jsonify({'error': 'One or more rows now conflict with an existing TR or email. Re-validate and try again.'}), 409

    log_tenant_audit(
        'faculty_bulk_import_users',
        target_type='user',
        description=f'Imported {imported_count} members from Faculty Excel import.',
        details={
            'imported_count': imported_count,
            'skipped_count': report['invalid_count'],
            'invalid_rows': report['invalid_rows'],
        },
        actor_user=user,
    )
    db.session.commit()
    _clear_faculty_dashboard_cache()
    return jsonify({
        'imported_count': imported_count,
        'skipped_count': report['invalid_count'],
        'invalid_rows': report['invalid_rows'],
    })


@faculty_bp.route('/faculty/meal-insights')
def meal_insights():
    user = _require_faculty()
    today = date.today()
    upcoming_menus = (
        tenant_filter(Menu.query)
        .filter(Menu.date >= today)
        .order_by(Menu.date.asc(), Menu.floor.asc(), Menu.meal_type.asc())
        .limit(60)
        .all()
    )
    historical_menus = (
        tenant_filter(Menu.query)
        .filter(Menu.date < today)
        .order_by(Menu.date.desc(), Menu.floor.asc(), Menu.meal_type.asc())
        .limit(160)
        .all()
    )

    historical_menu_ids = [menu.id for menu in historical_menus]
    feedback_map = {}
    if historical_menu_ids:
        feedback_map = {
            menu_id: {'avg_rating': round(float(avg_rating or 0), 1), 'feedback_count': int(count)}
            for menu_id, avg_rating, count in (
                tenant_filter(db.session.query(
                    Feedback.menu_id,
                    func.avg(Feedback.rating),
                    func.count(Feedback.id),
                ))
                .filter(Feedback.menu_id.in_(historical_menu_ids))
                .group_by(Feedback.menu_id)
                .all()
            )
        }

    dish_ids = set()
    for menu in upcoming_menus + historical_menus:
        if menu.dish_id:
            dish_ids.add(menu.dish_id)
        if menu.side_dish_id:
            dish_ids.add(menu.side_dish_id)

    suggestion_map = {}
    if dish_ids:
        suggestion_map = {
            dish_id: {'suggestion_count': int(suggestion_count), 'vote_count': int(vote_count)}
            for dish_id, suggestion_count, vote_count in (
                tenant_filter(db.session.query(
                    Suggestion.dish_id,
                    func.count(func.distinct(Suggestion.id)),
                    func.count(SuggestionVote.id),
                ))
                .outerjoin(SuggestionVote, SuggestionVote.suggestion_id == Suggestion.id)
                .filter(Suggestion.dish_id.in_(dish_ids))
                .group_by(Suggestion.dish_id)
                .all()
            )
        }

    def _suggestion_totals(menu):
        ids = {menu.dish_id, menu.side_dish_id}
        ids.discard(None)
        return {
            'suggestion_count': sum(suggestion_map.get(dish_id, {}).get('suggestion_count', 0) for dish_id in ids),
            'vote_count': sum(suggestion_map.get(dish_id, {}).get('vote_count', 0) for dish_id in ids),
        }

    history_rows = []
    for menu in historical_menus:
        feedback = feedback_map.get(menu.id, {'avg_rating': 0, 'feedback_count': 0})
        suggestions = _suggestion_totals(menu)
        history_rows.append({
            'menu': menu,
            'avg_rating': feedback['avg_rating'],
            'feedback_count': feedback['feedback_count'],
            'suggestion_count': suggestions['suggestion_count'],
            'vote_count': suggestions['vote_count'],
        })

    upcoming_rows = []
    for menu in upcoming_menus:
        suggestions = _suggestion_totals(menu)
        upcoming_rows.append({
            'menu': menu,
            'suggestion_count': suggestions['suggestion_count'],
            'vote_count': suggestions['vote_count'],
        })

    return render_template(
        'faculty/meal_insights.html',
        current_user=user,
        upcoming_rows=upcoming_rows,
        history_rows=history_rows,
        today=today,
    )


@faculty_bp.route('/faculty/profile', methods=['GET', 'POST'])
def profile():
    user = _require_faculty()

    if request.method == 'POST':
        user.full_name = (request.form.get('full_name') or '').strip() or None
        user.phone_number = (request.form.get('phone_number') or '').strip() or None
        _ensure_username_from_full_name(user, db.session)

        new_password = (request.form.get('new_password') or '').strip()
        if new_password:
            user.password_hash = generate_password_hash(new_password)

        db.session.commit()
        flash('Faculty profile updated successfully.', 'success')
        return redirect(url_for('faculty.profile'))

    return render_template(
        'faculty/profile.html',
        current_user=user,
        user=user,
    )


@faculty_bp.route('/faculty/messages', methods=['GET', 'POST'])
def messages():
    user = _require_faculty()
    floor_options = _get_tenant_floor_options(user)

    if request.method == 'POST':
        title = (request.form.get('title') or '').strip()
        content = (request.form.get('content') or '').strip()
        target_scope = (request.form.get('target_scope') or 'all_pantry_heads').strip()

        if not title or not content:
            flash('Message title and content are required.', 'error')
            return redirect(url_for('faculty.messages'))

        if target_scope not in {'all_pantry_heads', 'selected_floors'}:
            flash('Invalid recipient scope selected.', 'error')
            return redirect(url_for('faculty.messages'))

        selected_floors = []
        if target_scope == 'selected_floors':
            for raw_floor in request.form.getlist('target_floors'):
                try:
                    floor = int(raw_floor)
                except (TypeError, ValueError):
                    continue
                if floor in floor_options and floor not in selected_floors:
                    selected_floors.append(floor)

            if not selected_floors:
                flash('Choose at least one floor for a targeted Faculty message.', 'error')
                return redirect(url_for('faculty.messages'))

        message = FacultyMessage(
            title=title,
            content=content,
            target_scope=target_scope,
            created_by_id=user.id,
            tenant_id=getattr(g, 'tenant_id', None),
        )
        db.session.add(message)
        db.session.flush()

        if target_scope == 'selected_floors':
            for floor in selected_floors:
                db.session.add(FacultyMessageFloor(
                    faculty_message_id=message.id,
                    floor=floor,
                    tenant_id=getattr(g, 'tenant_id', None),
                ))

        recipient_query = tenant_filter(User.query).filter(User.role == 'pantryHead')
        if target_scope == 'selected_floors':
            recipient_query = recipient_query.filter(User.floor.in_(selected_floors))
        recipients = recipient_query.all()

        db.session.commit()

        if request.headers.get('Accept') == 'application/json':
            from flask import jsonify
            recipient_list = [{"id": r.id, "name": r.full_name or r.username or r.email or 'Pantry Head'} for r in recipients]
            return jsonify({"message_id": message.id, "recipients": recipient_list})

        flash('Faculty message created.', 'success')
        return redirect(url_for('faculty.messages'))

    active_messages = (
        tenant_filter(FacultyMessage.query)
        .filter_by(is_archived=False)
        .order_by(FacultyMessage.created_at.desc())
        .all()
    )
    archived_messages = (
        tenant_filter(FacultyMessage.query)
        .filter_by(is_archived=True)
        .order_by(FacultyMessage.created_at.desc())
        .all()
    )

    return render_template(
        'faculty/messages.html',
        current_user=user,
        active_messages=active_messages,
        archived_messages=archived_messages,
        floor_options=floor_options,
        message_target_floors=_message_target_floors,
    )


@faculty_bp.route('/faculty/messages/<int:message_id>/send_single', methods=['POST'])
def send_single_message(message_id):
    from flask import jsonify
    user = _require_faculty()
    message = tenant_filter(FacultyMessage.query).filter_by(id=message_id).first_or_404()
    
    data = request.get_json() or {}
    recipient_id = data.get('user_id')
    
    if not recipient_id:
        return jsonify({"status": "error", "message": "user_id is required"}), 400
        
    recipient = User.query.get(recipient_id)
    if not recipient:
        return jsonify({"status": "error", "message": "user not found"}), 404
        
    send_push_notification(
        user_id=recipient.id,
        title=f"Faculty Message: {message.title}",
        body=message.content[:100] + ("..." if len(message.content) > 100 else ""),
        icon="/static/icons/icon-192.png",
        url="/dashboard",
    )
    
    if recipient.email:
        tenant_name = getattr(g, 'tenant_name', 'Maskan')
        email_subject = f"[{tenant_name}] Faculty Message: {message.title}"
        email_html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <h2 style="color: #2c3e50;">{message.title}</h2>
            <p style="white-space: pre-wrap; font-size: 16px; color: #333;">{message.content}</p>
            <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
            <p style="font-size: 12px; color: #888;">Log in to your pantry dashboard to manage and reply if necessary.</p>
        </div>
        """
        send_email_notification(recipient.email, email_subject, email_html)
        
    return jsonify({"status": "success"})


@faculty_bp.route('/faculty/cycles', methods=['GET', 'POST'])
def cycles():
    user = _require_faculty()
    floor_options = _cycle_form_floor_options(user)

    if request.method == 'POST':
        action = (request.form.get('action') or 'save_draft').strip()
        cycle_data, error_message = _parse_cycle_form(floor_options)
        if error_message:
            flash(error_message, 'error')
            return redirect(url_for('faculty.cycles'))

        status = 'active' if action == 'activate_now' else 'draft'
        if status == 'active':
            existing_active = _current_active_cycle()
            if existing_active:
                flash('Close the current active cycle before activating a new one.', 'error')
                return redirect(url_for('faculty.cycles'))

        cycle = FacultyBudgetCycle(
            title=cycle_data['title'],
            start_date=cycle_data['start_date'],
            end_date=cycle_data['end_date'],
            submission_deadline=cycle_data['submission_deadline'],
            status=status,
            notes=cycle_data['notes'],
            created_by_id=user.id,
            activated_at=datetime.utcnow() if status == 'active' else None,
            tenant_id=getattr(g, 'tenant_id', None),
        )
        db.session.add(cycle)
        db.session.flush()

        _sync_cycle_budgets(cycle, user, cycle_data['allocations'])

        db.session.commit()
        flash(f'Faculty cycle "{cycle.title}" saved successfully.', 'success')
        return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))

    cycles = (
        tenant_filter(FacultyBudgetCycle.query)
        .order_by(FacultyBudgetCycle.created_at.desc())
        .all()
    )
    return render_template(
        'faculty/cycles.html',
        current_user=user,
        cycles=cycles,
        floor_options=floor_options,
        active_cycle=_current_active_cycle(),
        today=date.today(),
    )


@faculty_bp.route('/faculty/cycles/<int:cycle_id>')
def cycle_detail(cycle_id):
    user = _require_faculty()
    cycle = tenant_filter(FacultyBudgetCycle.query).filter_by(id=cycle_id).first_or_404()
    budgets = (
        tenant_filter(Budget.query)
        .filter_by(cycle_id=cycle.id)
        .order_by(Budget.floor.asc())
        .all()
    )
    submissions = (
        tenant_filter(FacultyReportSubmission.query)
        .filter_by(cycle_id=cycle.id)
        .all()
    )
    submissions_by_floor = {s.floor: s for s in submissions}
    floor_rows = []
    for budget in budgets:
        floor_rows.append({
            'floor': budget.floor,
            'budget': budget,
            'submission': submissions_by_floor.get(budget.floor),
        })

    editable_floors = _cycle_form_floor_options(user, cycle)
    budgets_by_floor = {budget.floor: budget for budget in budgets}

    active_cycle = _current_active_cycle()
    all_verified = _is_cycle_fully_verified(cycle)
    active_cycle_verified = _is_cycle_fully_verified(active_cycle) if active_cycle else True

    return render_template(
        'faculty/cycle_detail.html',
        current_user=user,
        cycle=cycle,
        floor_rows=floor_rows,
        editable_floors=editable_floors,
        budgets_by_floor=budgets_by_floor,
        today=date.today(),
        all_verified=all_verified,
        active_cycle=active_cycle,
        active_cycle_verified=active_cycle_verified,
    )


@faculty_bp.route('/faculty/cycles/<int:cycle_id>/edit', methods=['POST'])
def edit_cycle(cycle_id):
    user = _require_faculty()
    cycle = tenant_filter(FacultyBudgetCycle.query).filter_by(id=cycle_id).first_or_404()

    if cycle.status == 'closed':
        flash('Closed cycles cannot be edited.', 'error')
        return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))

    floor_options = _cycle_form_floor_options(user, cycle)
    cycle_data, error_message = _parse_cycle_form(floor_options)
    if error_message:
        flash(error_message, 'error')
        return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))

    cycle.title = cycle_data['title']
    cycle.start_date = cycle_data['start_date']
    cycle.end_date = cycle_data['end_date']
    cycle.submission_deadline = cycle_data['submission_deadline']
    cycle.notes = cycle_data['notes']
    _sync_cycle_budgets(cycle, user, cycle_data['allocations'])

    db.session.commit()
    flash('Cycle updated successfully.', 'success')
    return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))


@faculty_bp.route('/faculty/cycles/<int:cycle_id>/activate', methods=['POST'])
def activate_cycle(cycle_id):
    user = _require_faculty()
    cycle = tenant_filter(FacultyBudgetCycle.query).filter_by(id=cycle_id).first_or_404()
    existing_active = _current_active_cycle()
    if existing_active and existing_active.id != cycle.id:
        flash(f'Close the currently active cycle "{existing_active.title}" before activating another one.', 'error')
        return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))

    cycle.status = 'active'
    cycle.activated_at = datetime.utcnow()
    db.session.commit()
    flash('Cycle activated successfully.', 'success')
    return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))


@faculty_bp.route('/faculty/cycles/<int:cycle_id>/close', methods=['POST'])
def close_cycle(cycle_id):
    _require_faculty()
    cycle = tenant_filter(FacultyBudgetCycle.query).filter_by(id=cycle_id).first_or_404()
    if request.form.get('confirm_close') != '1':
        flash('Please confirm the close-cycle action before final closing.', 'error')
        return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))
    if cycle.status != 'active':
        flash('Only active cycles can be closed.', 'error')
        return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))
    if not _is_cycle_fully_verified(cycle):
        flash('Cannot close cycle until all floors have submitted reports and they are verified.', 'error')
        return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))
        
    cycle.status = 'closed'
    cycle.closed_at = datetime.utcnow()
    db.session.commit()
    flash('Cycle closed successfully.', 'success')
    return redirect(url_for('faculty.cycle_detail', cycle_id=cycle.id))


@faculty_bp.route('/faculty/cycles/<int:cycle_id>/delete', methods=['POST'])
def delete_cycle(cycle_id):
    _require_faculty()
    cycle = tenant_filter(FacultyBudgetCycle.query).filter_by(id=cycle_id).first_or_404()
    cycle_title = cycle.title
    _delete_cycle_related_data(cycle)
    db.session.commit()
    flash(f'Cycle "{cycle_title}" and all linked budgets, submissions, saved print reports, and PDFs were deleted.', 'success')
    return redirect(url_for('faculty.cycles'))


@faculty_bp.route('/reports', methods=['GET', 'POST'])
def reports_page():
    user = _require_user()
    if not user:
        flash('Your staff session expired. Please sign in again.', 'error')
        return redirect(url_for('auth.staff_login'))
    if user.role not in {'admin', 'pantryHead'}:
        abort(403)

    floor = _get_active_floor(user)
    faculty_workflow_enabled = current_tenant_faculty_workflow_enabled()
    active_cycle = _current_active_cycle()
    allocation = _cycle_for_floor(active_cycle.id, floor) if active_cycle else None
    submission = None
    saved_print_reports = []

    if faculty_workflow_enabled and active_cycle:
        submission = tenant_filter(FacultyReportSubmission.query).filter_by(
            cycle_id=active_cycle.id,
            floor=floor,
        ).first()
        floor_budget_ledger = build_floor_budget_ledger(
            floor=floor,
            faculty_workflow_enabled=faculty_workflow_enabled,
        )
        current_available_budget = floor_budget_ledger['current_available_budget']

        if request.method == 'POST':
            if not allocation:
                flash('This floor does not have an allocation in the active Faculty cycle.', 'error')
                return redirect(url_for('faculty.reports_page'))

            if submission and submission.status == 'verified':
                flash('This cycle report has already been verified.', 'error')
                return redirect(url_for('faculty.reports_page'))

            if submission and submission.status == 'submitted':
                flash('This floor already has a submitted report awaiting Faculty review.', 'error')
                return redirect(url_for('faculty.reports_page'))

            upload = request.files.get('report_pdf')
            submission_notes = (request.form.get('submission_notes') or '').strip()
            try:
                print_report_id = int(request.form.get('print_report_id') or 0)
            except (TypeError, ValueError):
                print_report_id = 0

            if not upload or not upload.filename:
                flash('Please upload the combined PDF report generated from Expenses.', 'error')
                return redirect(url_for('faculty.reports_page'))

            if not upload.filename.lower().endswith('.pdf'):
                flash('Only PDF uploads are allowed for Faculty submissions.', 'error')
                return redirect(url_for('faculty.reports_page'))

            if not print_report_id:
                flash('Please choose a saved print report from Expenses first.', 'error')
                return redirect(url_for('faculty.reports_page'))

            selected_print_report = tenant_filter(ExpensePrintReport.query).filter_by(
                id=print_report_id,
                floor=floor,
            ).first()
            if not selected_print_report:
                flash('The selected print report was not found for this floor.', 'error')
                return redirect(url_for('faculty.reports_page'))
            if selected_print_report.cycle_id and selected_print_report.cycle_id != active_cycle.id:
                flash('Please choose a print report created for the current active cycle.', 'error')
                return redirect(url_for('faculty.reports_page'))

            selected_bills = []
            for link in selected_print_report.bill_links:
                if link.bill and link.bill.floor == floor:
                    selected_bills.append(link.bill)

            if not selected_bills:
                flash('The selected print report does not contain any bills.', 'error')
                return redirect(url_for('faculty.reports_page'))

            for bill in selected_bills:
                if bill.report_submission_id and (not submission or bill.report_submission_id != submission.id):
                    flash(f'Bill {bill.bill_no} is already linked to another report submission.', 'error')
                    return redirect(url_for('faculty.reports_page'))

            revision_no = (submission.revision_no + 1) if submission else 1
            stored_filename, storage_path, file_size = _save_submission_file(upload, active_cycle, floor, revision_no)

            if submission:
                old_path = submission.storage_path
                submission.report_title = selected_print_report.report_title
                submission.print_report_id = selected_print_report.id
                submission.status = 'submitted'
                submission.allocated_amount = current_available_budget
                submission.submission_notes = submission_notes
                submission.review_notes = None
                submission.stored_filename = stored_filename
                submission.original_filename = upload.filename
                submission.storage_path = storage_path
                submission.file_size_bytes = file_size
                submission.revision_no = revision_no
                submission.uploaded_by_id = user.id
                submission.submitted_at = datetime.utcnow()
                submission.verified_at = None
                submission.verified_by_id = None
                _sync_submission_links(submission, selected_bills, [])
                if old_path and old_path != storage_path and os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except OSError:
                        pass
            else:
                submission = FacultyReportSubmission(
                    cycle_id=active_cycle.id,
                    print_report_id=selected_print_report.id,
                    floor=floor,
                    uploaded_by_id=user.id,
                    report_title=selected_print_report.report_title,
                    status='submitted',
                    allocated_amount=current_available_budget,
                    submission_notes=submission_notes,
                    stored_filename=stored_filename,
                    original_filename=upload.filename,
                    storage_path=storage_path,
                    file_size_bytes=file_size,
                    revision_no=revision_no,
                    submitted_at=datetime.utcnow(),
                    tenant_id=getattr(g, 'tenant_id', None),
                )
                db.session.add(submission)
                db.session.flush()
                _sync_submission_links(submission, selected_bills, [])

            db.session.commit()
            flash('Report submitted to Faculty successfully.', 'success')
            return redirect(url_for('faculty.reports_page'))

        saved_print_reports = _saved_print_reports_for_floor(floor, active_cycle.id)

    # Fetch Ad-hoc / irregular saved reports for this floor
    adhoc_reports = tenant_filter(ExpensePrintReport.query).filter(
        ExpensePrintReport.floor == floor,
        ExpensePrintReport.cycle_id == None,
        ExpensePrintReport.storage_path != None
    ).order_by(ExpensePrintReport.created_at.desc()).all()

    return render_template(
        'reports.html',
        current_user=user,
        active_floor=floor,
        faculty_workflow_enabled=faculty_workflow_enabled,
        active_cycle=active_cycle,
        allocation=allocation,
        submission=submission,
        saved_print_reports=saved_print_reports,
        adhoc_reports=adhoc_reports,
        today=date.today(),
    )

from flask import send_from_directory

@faculty_bp.route('/reports/adhoc/<int:report_id>/download')
def download_adhoc_report(report_id):
    user = _require_user()
    if not user or user.role not in ['admin', 'pantryHead']:
        abort(403)
        
    floor = _get_active_floor(user)
    report = tenant_filter(ExpensePrintReport.query).filter_by(id=report_id).first_or_404()
    
    if user.role != 'admin' and report.floor != floor:
        abort(403)
        
    if not report.storage_path:
        flash("This report does not have a saved PDF file.", "error")
        return redirect(url_for('faculty.reports_page'))
        
    base_dir = current_app.config.get("REPORT_STORAGE_ROOT", os.path.join(current_app.root_path, "tmp", "reports"))
    
    file_dir = os.path.dirname(os.path.normpath(os.path.join(base_dir, report.storage_path)))
    filename = os.path.basename(report.storage_path)
    
    try:
        return send_from_directory(file_dir, filename, as_attachment=True, download_name=report.stored_filename)
    except FileNotFoundError:
        flash("File missing from server. It might have been deleted.", "error")
        return redirect(url_for('faculty.reports_page'))

@faculty_bp.route('/reports/adhoc/<int:report_id>/delete', methods=['POST'])
def delete_adhoc_report(report_id):
    user = _require_user()
    if not user or user.role not in ['admin', 'pantryHead']:
        abort(403)
        
    floor = _get_active_floor(user)
    report = tenant_filter(ExpensePrintReport.query).filter_by(id=report_id).first_or_404()
    
    if user.role != 'admin' and report.floor != floor:
        abort(403)

    if report.storage_path:
        base_dir = current_app.config.get("REPORT_STORAGE_ROOT", os.path.join(current_app.root_path, "tmp", "reports"))
        abs_path = os.path.normpath(os.path.join(base_dir, report.storage_path))
        if os.path.exists(abs_path):
            try:
                os.remove(abs_path)
            except OSError:
                pass
                
    db.session.delete(report)
    db.session.commit()
    
    flash("Ad-Hoc Report and its PDF have been permanently deleted.", "success")
    return redirect(url_for('faculty.reports_page'))

@faculty_bp.route('/faculty/reports/<int:submission_id>')
def report_detail(submission_id):
    user = _require_faculty()
    submission = tenant_filter(FacultyReportSubmission.query).filter_by(id=submission_id).first_or_404()
    linked_bills = (
        tenant_filter(Bill.query)
        .filter_by(report_submission_id=submission.id)
        .order_by(Bill.bill_date.desc(), Bill.created_at.desc())
        .all()
    )
    verification_summary = _build_submission_verification_data(submission, linked_bills)
    return render_template(
        'faculty/report_detail.html',
        current_user=user,
        submission=submission,
        cycle=submission.cycle,
        linked_bills=linked_bills,
        verification_summary=verification_summary,
    )


@faculty_bp.route('/faculty/reports/<int:submission_id>/verify', methods=['POST'])
def verify_report(submission_id):
    user = _require_faculty()
    submission = tenant_filter(FacultyReportSubmission.query).filter_by(id=submission_id).first_or_404()
    if request.form.get('verification_acknowledged') != '1':
        flash('Please review the verification modal before approving this report.', 'error')
        return redirect(url_for('faculty.report_detail', submission_id=submission.id))
    submission.status = 'verified'
    submission.review_notes = (request.form.get('review_notes') or '').strip()
    submission.verified_at = datetime.utcnow()
    submission.verified_by_id = user.id
    db.session.commit()
    flash('Report verified successfully.', 'success')
    return redirect(url_for('faculty.report_detail', submission_id=submission.id))


@faculty_bp.route('/faculty/reports/<int:submission_id>/reject', methods=['POST'])
def reject_report(submission_id):
    _require_faculty()
    submission = tenant_filter(FacultyReportSubmission.query).filter_by(id=submission_id).first_or_404()
    review_notes = (request.form.get('review_notes') or '').strip()
    if not review_notes:
        flash('Please add review notes before rejecting a report.', 'error')
        return redirect(url_for('faculty.report_detail', submission_id=submission.id))

    submission.status = 'rejected'
    submission.review_notes = review_notes
    submission.verified_at = None
    submission.verified_by_id = None
    db.session.commit()
    flash('Report rejected. The floor can now revise and re-submit.', 'success')
    return redirect(url_for('faculty.report_detail', submission_id=submission.id))


@faculty_bp.route('/faculty/messages/<int:message_id>/archive', methods=['POST'])
def archive_message(message_id):
    _require_faculty()
    message = tenant_filter(FacultyMessage.query).filter_by(id=message_id).first_or_404()
    message.is_archived = True
    db.session.commit()
    flash('Faculty message archived.', 'success')
    return redirect(url_for('faculty.messages'))


@faculty_bp.route('/faculty/messages/<int:message_id>/delete', methods=['POST'])
def delete_message(message_id):
    _require_faculty()
    message = tenant_filter(FacultyMessage.query).filter_by(id=message_id).first_or_404()
    db.session.delete(message)
    db.session.commit()
    flash('Faculty message deleted.', 'success')
    return redirect(url_for('faculty.messages'))


@faculty_bp.route('/faculty/reports/<int:submission_id>/download')
def download_report(submission_id):
    _require_faculty()
    if not current_tenant_faculty_workflow_enabled():
        flash('Faculty workflow is disabled for this tenant right now.', 'error')
        return redirect(url_for('faculty.login'))
    submission = tenant_filter(FacultyReportSubmission.query).filter_by(id=submission_id).first_or_404()
    
    if not submission.storage_path:
        flash('No storage path found for this report.', 'error')
        return redirect(url_for('faculty.report_detail', submission_id=submission.id))

    if os.path.isabs(submission.storage_path):
        abs_path = submission.storage_path
    else:
        abs_path = os.path.normpath(os.path.join(current_app.config["REPORT_STORAGE_ROOT"], submission.storage_path))

    if not os.path.exists(abs_path):
        flash('The stored PDF could not be found on the server.', 'error')
        return redirect(url_for('faculty.report_detail', submission_id=submission.id))

    return send_file(
        abs_path,
        as_attachment=True,
        download_name=submission.stored_filename,
        mimetype='application/pdf',
    )


@faculty_bp.route('/reports/<int:submission_id>/download')
def download_floor_submission(submission_id):
    user = _require_user()
    if not user:
        flash('Your staff session expired. Please sign in again.', 'error')
        return redirect(url_for('auth.staff_login'))
    if user.role not in {'admin', 'pantryHead'}:
        abort(403)
    if not current_tenant_faculty_workflow_enabled():
        flash('Faculty report submission is disabled for this tenant. Use Expenses for manual budgeting and printing.', 'error')
        return redirect(url_for('finance.expenses'))

    submission = tenant_filter(FacultyReportSubmission.query).filter_by(id=submission_id).first_or_404()
    floor = _get_active_floor(user)
    if submission.floor != floor:
        abort(403)

    if not submission.storage_path:
        flash('No storage path found for this report.', 'error')
        return redirect(url_for('faculty.reports_page'))

    if os.path.isabs(submission.storage_path):
        abs_path = submission.storage_path
    else:
        abs_path = os.path.normpath(os.path.join(current_app.config["REPORT_STORAGE_ROOT"], submission.storage_path))

    if not os.path.exists(abs_path):
        flash('The stored PDF could not be found on the server.', 'error')
        return redirect(url_for('faculty.reports_page'))

    return send_file(
        abs_path,
        as_attachment=True,
        download_name=submission.stored_filename,
        mimetype='application/pdf',
    )
